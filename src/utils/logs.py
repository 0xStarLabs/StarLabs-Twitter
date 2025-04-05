import os
from typing import Dict, Any, Optional, List, Tuple
from asyncio import Lock
from tqdm import tqdm
from dataclasses import dataclass
import asyncio
import random
from loguru import logger
import threading
import openpyxl
from openpyxl.styles import PatternFill
from dataclasses import dataclass
import pandas as pd
from datetime import datetime
from rich.progress import (
    Progress,
    SpinnerColumn,
    TextColumn,
    BarColumn,
    TimeElapsedColumn,
)
from collections import Counter


# Global mutex for Excel file access
excel_mutex = threading.Lock()
excel_mutex_async = asyncio.Lock()

# Thread-safe dictionary to store task logs
task_logs: Dict[str, List[Tuple[int, str, str]]] = {}
task_logs_lock = asyncio.Lock()


@dataclass
class ProgressTracker:
    total: int
    current: int = 0
    description: str = "Progress"
    _lock: Lock = Lock()
    bar_length: int = 30  # Длина прогресс-бара в символах

    def __post_init__(self):
        pass

    def _create_progress_bar(self, percentage: float) -> str:
        filled_length = int(self.bar_length * percentage / 100)
        bar = "█" * filled_length + "░" * (self.bar_length - filled_length)
        return bar

    async def increment(self, amount: int = 1, message: Optional[str] = None):
        async with self._lock:
            self.current += amount
            percentage = (self.current / self.total) * 100
            bar = self._create_progress_bar(percentage)

            # Добавляем эмодзи в зависимости от прогресса
            emoji = "⏳"
            if percentage >= 100:
                emoji = "✅"
            elif percentage >= 50:
                emoji = "🔄"

            progress_msg = f"{emoji} [{self.description}] [{bar}] {self.current}/{self.total} ({percentage:.1f}%)"
            # if message:
            #     progress_msg += f"\n    ├─ {message}"
            logger.info(progress_msg)

    async def set_total(self, total: int):
        async with self._lock:
            self.total = total

    def __del__(self):
        pass  # Убираем закрытие tqdm


async def create_progress_tracker(
    total: int, description: str = "Progress"
) -> ProgressTracker:
    return ProgressTracker(total=total, description=description)


async def process_item(tracker: ProgressTracker, item_id: int):
    delay = random.uniform(2, 5)
    await asyncio.sleep(delay)
    status = "completed" if random.random() > 0.2 else "pending"
    await tracker.increment(1, f"📝 Account {item_id} status: {status}")


async def update_account_in_excel(
    file_path: str, auth_token: str, username: str = None, status: str = None
) -> None:
    """
    Update account fields in Excel file for specified auth_token.

    Args:
        file_path: Path to the Excel file
        auth_token: Authentication token to identify the account
        username: New username to set (optional)
        status: New status to set (optional)

    Raises:
        Exception: If any error occurs during the update process
    """
    # Lock the mutex before accessing the file
    async with excel_mutex_async:
        try:
            # Run the blocking file operations in a thread pool
            def update_excel():
                with excel_mutex:
                    # Open the Excel file
                    workbook = openpyxl.load_workbook(file_path)

                    # Get the first sheet
                    sheet = workbook.active

                    # Find the row with matching auth token
                    row_index = None
                    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
                        if row[0] == auth_token:
                            row_index = i
                            break

                    if row_index is None:
                        raise Exception("Account not found in Excel file")

                    updated_fields = []

                    # Update username if provided
                    if username is not None:
                        cell = sheet.cell(row=row_index, column=3)  # Column C
                        cell.value = username
                        updated_fields.append(f"username={username}")

                    # Update status if provided
                    if status is not None:
                        cell = sheet.cell(row=row_index, column=4)  # Column D
                        cell.value = status
                        updated_fields.append(f"status={status}")

                        cell.fill = PatternFill(
                            start_color="ADD8E6",
                            end_color="ADD8E6",
                            fill_type="solid",
                        )  # Light blue

                    # Save the changes if any updates were made
                    if updated_fields:
                        workbook.save(file_path)

            # Run the blocking operation in a thread pool
            await asyncio.to_thread(update_excel)

        except Exception as e:
            error_msg = f"Failed to update Excel file: {str(e)}"
            logger.error(error_msg)
            raise Exception(error_msg)


async def add_task_log_entry(
    task_name: str, account_index: int, auth_token: str, status: str
):
    """
    Add a log entry to the in-memory task log dictionary in a thread-safe manner.

    Args:
        task_name: Name of the task (like, retweet, etc.)
        account_index: Index of the account
        auth_token: Auth token of the account
        status: Status of the task (success, failed, etc.)
    """
    async with task_logs_lock:
        if task_name not in task_logs:
            task_logs[task_name] = []

        # Add entry to the in-memory log
        task_logs[task_name].append((account_index, auth_token, status))


async def get_task_logs_summary():
    """
    Get a summary of all task logs for reporting.

    Returns:
        Dict containing:
        - tasks: List of task names
        - total_accounts: Total number of accounts processed
        - success_count: Number of successful task executions
        - total_tasks: Total number of task executions
        - task_success_counts: Dictionary with counts of successful executions by task
        - task_counts: Dictionary with total counts by task
    """
    async with task_logs_lock:
        summary = {
            "tasks": list(task_logs.keys()),
            "total_accounts": set(),
            "success_count": 0,
            "total_tasks": 0,
            "task_success_counts": {},
            "task_counts": {},
        }

        for task_name, entries in task_logs.items():
            # Count successful tasks for this task type
            task_success = sum(1 for _, _, status in entries if status == "success")
            summary["task_success_counts"][task_name] = task_success
            summary["task_counts"][task_name] = len(entries)

            # Add to total counts
            summary["total_tasks"] += len(entries)
            summary["success_count"] += task_success

            # Add unique accounts
            for _, auth_token, _ in entries:
                summary["total_accounts"].add(auth_token)

        # Convert set to count
        summary["total_accounts"] = len(summary["total_accounts"])

        return summary


async def save_task_logs_to_excel():
    """
    Save all collected task logs to Excel files.
    This should be called at the end of processing.

    Returns:
        Dict: A summary of task execution statistics
    """
    try:
        # Ensure logs directory exists
        os.makedirs("logs", exist_ok=True)

        # Get current UTC timestamp for filenames
        timestamp = datetime.utcnow().strftime("%Y_%m_%d_%Hx%M")

        # Get summary before saving
        summary = await get_task_logs_summary()

        async with task_logs_lock:
            # Process each task's logs
            for task_name, entries in task_logs.items():
                # Create filename with task name and timestamp
                filename = f"logs/{task_name}_{timestamp}.xlsx"

                # Sort entries by account index
                entries.sort(key=lambda x: x[0])

                # Create DataFrame for this task
                df = pd.DataFrame(
                    entries, columns=["account_index", "auth_token", "status"]
                )

                # Save to Excel
                df.to_excel(filename, index=False)
                logger.success(f"Task log saved to {filename}")

            # Clear the logs after saving
            task_logs.clear()

        return summary

    except Exception as e:
        logger.error(f"Failed to save task logs to Excel: {e}")
        return None
